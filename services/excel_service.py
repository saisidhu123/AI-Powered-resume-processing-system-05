import os
from typing import List, Dict, Any, Tuple
import openpyxl
import pandas as pd

def read_excel_headers(file_path: str) -> Tuple[bool, List[str], str]:
    """
    Read top row as Excel headers from uploaded XLSX file.
    Preserves exact column header names and sequence.
    """
    if not os.path.exists(file_path):
        return False, [], "Excel file does not exist on disk."

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        
        headers = []
        for cell in sheet[1]: # Row 1
            val = str(cell.value or "").strip()
            if val:
                headers.append(val)

        wb.close()

        if not headers:
            return False, [], "The uploaded Excel file does not contain any valid column headers in Row 1."

        return True, headers, ""
    except Exception as e:
        return False, [], f"Failed to read Excel headers: {str(e)}"


def read_existing_candidate_rows(file_path: str) -> List[Dict[str, Any]]:
    """
    Read existing candidate records from an Excel template or previous output file.
    Returns list of candidate dictionary rows.
    """
    if not os.path.exists(file_path):
        return []

    try:
        df = pd.read_excel(file_path)
        if df.empty:
            return []
        
        # Clean columns
        df.columns = [str(col).strip() for col in df.columns]
        
        records = df.to_dict(orient="records")
        # Clean NaNs to empty string
        cleaned_records = []
        for r in records:
            clean_r = {}
            for k, v in r.items():
                if pd.isna(v):
                    clean_r[k] = ""
                else:
                    clean_r[k] = str(v).strip()
            cleaned_records.append(clean_r)

        return cleaned_records
    except Exception:
        return []


import zipfile
import re
from utils.helpers import SUPPORTED_DOMAINS

def is_tech_domain_header(header_name: str) -> bool:
    """
    Check if a header string represents a technology domain concept.
    """
    if not header_name:
        return False
    h = str(header_name).strip().lower().replace("_", " ").replace("-", " ")
    h = re.sub(r"\s+", " ", h).strip()
    return any(k in h for k in [
        "technology domain", "tech domain", "domain", "technology category",
        "tech category", "tech stack", "technology"
    ]) and not any(k in h for k in ["skill", "tools", "competenc"])


def sanitize_sheet_name(name: str) -> str:
    r"""
    Sanitize sheet name to conform with Excel requirements:
    - Max 31 characters
    - Cannot contain: \ / ? * : [ ]
    """
    cleaned = str(name).replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace(":", "-").replace("[", "(").replace("]", ")").strip()
    return cleaned[:31] if cleaned else "Sheet"


def populate_excel_template(template_path: str, candidate_data: Dict[str, Any], output_path: str) -> Tuple[bool, str]:
    """
    Populate Excel template with a single new candidate row.
    - Preserves uploaded template's column headers
    - Preserves column order
    - Maps Technology Domain to existing matching column or appends 'Technology Domain'
    - Appends candidate row
    - Leaves empty values blank
    - Saves output XLSX file
    """
    if not os.path.exists(template_path):
        return False, f"Template path '{template_path}' does not exist."

    try:
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active

        # Read top row headers to determine column indices
        header_col_map = {}
        has_tech_col = False
        tech_col_idx = None

        for col_idx, cell in enumerate(sheet[1], start=1):
            h_name = str(cell.value or "").strip()
            if h_name:
                header_col_map[col_idx] = h_name
                if is_tech_domain_header(h_name):
                    has_tech_col = True
                    tech_col_idx = col_idx

        if not header_col_map:
            return False, "No valid headers found in row 1 of template."

        # If template does not have a Technology Domain column, append one
        if not has_tech_col:
            next_col_idx = len(header_col_map) + 1
            header_col_map[next_col_idx] = "Technology Domain"
            sheet.cell(row=1, column=next_col_idx, value="Technology Domain")
            tech_col_idx = next_col_idx

        # Find the next empty row
        next_row = sheet.max_row + 1
        if sheet.max_row == 1:
            next_row = 2

        tech_domain_val = candidate_data.get("Technology Domain") or ", ".join(candidate_data.get("_tech_domains", []))

        # Populate cell for each mapped column header
        for col_idx, h_name in header_col_map.items():
            if col_idx == tech_col_idx:
                value = tech_domain_val
            else:
                value = candidate_data.get(h_name, "")
            
            cell = sheet.cell(row=next_row, column=col_idx)
            if value != "" and value is not None:
                cell.value = str(value)
            else:
                cell.value = None

        # Ensure output directory exists
        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        wb.save(output_path)
        wb.close()

        return True, ""
    except Exception as e:
        return False, f"Error populating Excel template: {str(e)}"


def populate_excel_template_batch(template_path: str, candidates: List[Dict[str, Any]], output_path: str) -> Tuple[bool, str]:
    """
    Populate Excel template with multiple candidate rows and multi-sheet domain breakdown:
    - Sheet 'All Candidates': Complete processed candidate dataset
    - Technology-specific Sheets: Non-empty sheets for each detected domain (e.g. 'Java', 'Python', 'AI-ML', etc.)
    - Appends 'Technology Domain' column if not present in template without duplicating columns.
    """
    if not os.path.exists(template_path):
        return False, f"Template path '{template_path}' does not exist."

    try:
        wb = openpyxl.load_workbook(template_path)
        sheet = wb.active
        sheet.title = "All Candidates"

        header_col_map = {}
        has_tech_col = False
        tech_col_idx = None

        for col_idx, cell in enumerate(sheet[1], start=1):
            h_name = str(cell.value or "").strip()
            if h_name:
                header_col_map[col_idx] = h_name
                if is_tech_domain_header(h_name):
                    has_tech_col = True
                    tech_col_idx = col_idx

        if not header_col_map:
            return False, "No valid headers found in row 1 of template."

        # If template does not have a Technology Domain column, add it
        if not has_tech_col:
            next_col_idx = len(header_col_map) + 1
            header_col_map[next_col_idx] = "Technology Domain"
            sheet.cell(row=1, column=next_col_idx, value="Technology Domain")
            tech_col_idx = next_col_idx

        # 1. Populate 'All Candidates' sheet
        start_row = 2
        for idx, candidate in enumerate(candidates):
            current_row = start_row + idx
            tech_val = candidate.get("Technology Domain") or ", ".join(candidate.get("_tech_domains", []))
            for col_idx, h_name in header_col_map.items():
                if col_idx == tech_col_idx:
                    value = tech_val
                else:
                    value = candidate.get(h_name, "")

                cell = sheet.cell(row=current_row, column=col_idx)
                if value != "" and value is not None:
                    cell.value = str(value)
                else:
                    cell.value = None

        # 2. Generate Technology-Specific Sheets for active domains
        for domain in SUPPORTED_DOMAINS:
            # Match candidates belonging to this domain
            matching_candidates = []
            for c in candidates:
                cand_domains = c.get("_tech_domains", [])
                cand_domain_str = str(c.get("Technology Domain", ""))
                if domain in cand_domains or domain.lower() in cand_domain_str.lower():
                    matching_candidates.append(c)

            # Skip empty technology sheets
            if not matching_candidates:
                continue

            sheet_title = sanitize_sheet_name(domain)
            # Create tech sheet
            tech_sheet = wb.create_sheet(title=sheet_title)

            # Write header row
            for col_idx, h_name in header_col_map.items():
                tech_sheet.cell(row=1, column=col_idx, value=h_name)

            # Write matching candidate rows
            for c_idx, candidate in enumerate(matching_candidates):
                c_row = start_row + c_idx
                tech_val = candidate.get("Technology Domain") or ", ".join(candidate.get("_tech_domains", []))
                for col_idx, h_name in header_col_map.items():
                    if col_idx == tech_col_idx:
                        value = tech_val
                    else:
                        value = candidate.get(h_name, "")

                    cell = tech_sheet.cell(row=c_row, column=col_idx)
                    if value != "" and value is not None:
                        cell.value = str(value)
                    else:
                        cell.value = None

        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        wb.save(output_path)
        wb.close()
        return True, ""
    except Exception as e:
        return False, f"Error generating Master Excel batch file: {str(e)}"


def generate_duplicate_report(duplicates: List[Dict[str, Any]], excel_headers: List[str], output_path: str) -> Tuple[bool, str]:
    """
    Generate Duplicate Candidate Report (.xlsx).
    """
    try:
        rows = []
        for d in duplicates:
            r = {
                "Filename": d.get("_filename", "Unknown"),
                "Duplicate Reason": d.get("_duplicate_reason", "Matched Identifier")
            }
            for h in excel_headers:
                r[h] = d.get(h, "")
            rows.append(r)

        df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=["Filename", "Duplicate Reason"] + excel_headers)
        
        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        df.to_excel(output_path, index=False)
        return True, ""
    except Exception as e:
        return False, f"Error generating Duplicate Report: {str(e)}"


def generate_error_and_missing_report(failed_resumes: List[Dict[str, Any]], missing_fields_log: List[Dict[str, Any]], output_path: str) -> Tuple[bool, str]:
    """
    Generate Error and Missing Fields Report (.xlsx) with two sheets.
    """
    try:
        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_failed = pd.DataFrame(failed_resumes) if failed_resumes else pd.DataFrame(columns=["Filename", "Error"])
            df_failed.to_excel(writer, sheet_name="Failed Resumes", index=False)

            df_missing = pd.DataFrame(missing_fields_log) if missing_fields_log else pd.DataFrame(columns=["Filename", "Candidate Name", "Missing Fields Count", "Missing Field Names"])
            df_missing.to_excel(writer, sheet_name="Missing Fields Log", index=False)

        return True, ""
    except Exception as e:
        return False, f"Error generating Error & Missing Fields Report: {str(e)}"


def generate_classification_report(tech_stats: Dict[str, int], exp_stats: Dict[str, int], output_path: str) -> Tuple[bool, str]:
    """
    Generate Classification Analytics Summary Report (.xlsx).
    """
    try:
        out_dir = os.path.dirname(output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        df_tech = pd.DataFrame([{"Technology Domain": k, "Candidate Count": v} for k, v in tech_stats.items()])
        df_exp = pd.DataFrame([{"Experience Category": k, "Candidate Count": v} for k, v in exp_stats.items()])

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df_tech.to_excel(writer, sheet_name="Technology Breakdown", index=False)
            df_exp.to_excel(writer, sheet_name="Experience Breakdown", index=False)

        return True, ""
    except Exception as e:
        return False, f"Error generating Classification Report: {str(e)}"


def create_batch_zip_package(report_files: List[Tuple[str, str]], zip_output_path: str) -> Tuple[bool, str]:
    """
    Bundle multiple generated report files into a single ZIP package.
    report_files is a list of tuples: [(file_path, arcname_inside_zip), ...]
    """
    try:
        out_dir = os.path.dirname(zip_output_path)
        if out_dir and not os.path.exists(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        with zipfile.ZipFile(zip_output_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for fpath, arcname in report_files:
                if os.path.exists(fpath):
                    zf.write(fpath, arcname=arcname)

        return True, ""
    except Exception as e:
        return False, f"Error creating ZIP package: {str(e)}"
